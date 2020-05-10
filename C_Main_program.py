from C_install_libraries import *
installIfNeeded("numpy")
installIfNeeded("pillow")
installIfNeeded("openpyxl")
installIfNeeded("matplotlib")
installIfNeeded("pandas")
installIfNeeded("yfinance")
installIfNeeded("scipy")
installIfNeeded("statsmodels")

from tkinter import *
from C_indicators import *
import numpy as np
from C_download_data import *
from C_ARIMA import text_ARIMA, predict_ARIMA
from PIL import Image, ImageTk
from tkinter import messagebox
from openpyxl import load_workbook

global name_ind_param, index_num, help_num
index_num = np.nan
names = ["S&P 500", "SSE", "Euronext 100", "Hang Seng", "Kospi", "Nikkei 225", "BSE"]
TA_name = ["SMA", "EMA", "LWMA", "Aroon", "CCI", "SO", "CMO", "MAE(SMA, LL)", "MAE(SMA, UL)", "MAE(LWMA, LL)", "MAE(LWMA, UL)", "MAE(EMA, LL)", "MAE(EMA, UL)"]

class Invest_model(Frame):
	def __init__(self, parent):
		Frame.__init__(self, parent)
		self.parent = parent
		self.setUI()

	def setUI(self):
		self.parent.title("Бот для торгівлі на фондовому ринку")
		self.parent.iconbitmap(r'USD.ico')
		self.pack(fill = BOTH, expand=1)
		self.centerWindow()

		self.label0 = Label(self, text='Вітаю!', font=('Arial', 20, 'bold'), fg='red').place(x=100, y=10)
		self.label1 = Label(self, text='--- Оберіть індекс ---', font=('Arial', 15)).place(x=55, y=55)
		self.index0 = Button(self, text="S&P 500", font = ('Arial', 15, 'bold'), activebackground='blue', activeforeground='white',\
		 fg='blue', bg='white', width=15, relief=GROOVE, command=lambda: self.analysis(0)).place(x=59, y=95)
		self.index0 = Button(self, text="SSE", font = ('Arial', 15, 'bold'), activebackground='blue', activeforeground='white',\
		 fg='blue', bg='white', width=15, relief=GROOVE, command=lambda: self.analysis(1)).place(x=59, y=140)
		self.index0 = Button(self, text="Euronext 100", font = ('Arial', 15, 'bold'), activebackground='blue', activeforeground='white',\
		 fg='blue', bg='white', width=15, relief=GROOVE, command=lambda: self.analysis(2)).place(x=59, y=185)
		self.index0 = Button(self, text="Hang Seng", font = ('Arial', 15, 'bold'), activebackground='blue', activeforeground='white',\
		 fg='blue', bg='white', width=15, relief=GROOVE, command=lambda: self.analysis(3)).place(x=59, y=230)
		self.index0 = Button(self, text="Kospi", font = ('Arial', 15, 'bold'), activebackground='blue', activeforeground='white',\
		 fg='blue', bg='white', width=15, relief=GROOVE, command=lambda: self.analysis(4)).place(x=59, y=275)
		self.index0 = Button(self, text="Nikkei 225", font = ('Arial', 15, 'bold'), activebackground='blue', activeforeground='white',\
		 fg='blue', bg='white', width=15, relief=GROOVE, command=lambda: self.analysis(5)).place(x=59, y=320)
		self.index0 = Button(self, text="BSE", font = ('Arial', 15, 'bold'), activebackground='blue', activeforeground='white',\
		 fg='blue', bg='white', width=15, relief=GROOVE, command=lambda: self.analysis(6)).place(x=59, y=365)

	def centerWindow(self):
		w = 300
		h = 425
		x = (1536 - w)/2
		y = (864 - h)/2
		self.parent.geometry('%dx%d+%d+%d' % (w, h, x, y))

	def analysis(self, num):
		global index_num
		index_num = num
		window = Toplevel(self.parent)
		my_window = predict_window(window)

class predict_window():
	def __init__(self, parent):
		global index_num
		self.parent = parent
		self.parent.grab_set()

		self.parent.iconbitmap(r'USD.ico')
		self.centerWindow()
		self.parent.resizable(False, False)

		self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 3)).pack()
		self.label0 = Label(self.parent, text=names[index_num], anchor=N, font=('Arial', 20, 'bold'), fg='red').pack()

		self.label = Label(self.parent, text = '--- Значення індексу ---', font=('Arial', 14, 'bold'), fg='blue').place(x=25, y=135)

		for i in range(len(date[index_num])):
			self.label = Label(self.parent, text=date[index_num][i], font=('Arial', 13)).place(x=50, y=170+i*25)
			len_index = len(index_list[index_num]['Close'])
			self.label = Label(self.parent, text='{0:.2f}'.format(round(index_list[index_num]['Close'][len_index-i-1], 2)),\
			 font=('Arial', 13)).place(x=145, y=170+i*25)

		self.label = Label(self.parent, text = '--- Сигнали ---', font=('Arial', 14, 'bold'), fg='blue').place(x=305, y=55)
		if text_ARIMA[index_num] != 'ARIMA(nan, nan, nan, nan)':
			if predict_ARIMA[index_num] == 1:
				color = 'green'
			elif predict_ARIMA[index_num] == -1:
				color = 'orangered'
			else:
				color = 'grey'
			self.label = Label(self.parent, text=text_ARIMA[index_num], font=('Arial', 13)).place(x=270, y=90)
			self.label = Label(self.parent, text=predict_ARIMA[index_num], font=('Arial', 13, 'bold'), fg=color).place(x=420, y=90)
		else:
			self.label = Label(self.parent, text="ARIMA",  font=('Arial', 13)).place(x=270, y=90)
			self.label = Label(self.parent, text='---', font=('Arial', 13, 'bold'), fg='crimson').place(x=420, y=90)

		for i in range(len(TA_name)):
			self.label = Label(self.parent, text=TA_name[i], font=('Arial', 13)).place(x=270, y=115+i*25)
			if list_TA_predict[index_num][i] == 1:
				self.label = Label(self.parent, text=list_TA_predict[index_num][i], font=('Arial', 13, 'bold'), fg='green').place(x=420, y=115+i*25)
			elif list_TA_predict[index_num][i] == -1:
				self.label = Label(self.parent, text=list_TA_predict[index_num][i], font=('Arial', 13, 'bold'), fg='orangered').place(x=420, y=115+i*25)
			elif list_TA_predict[index_num][i] == "nan":
				self.label = Label(self.parent, text="---", font=('Arial', 13, 'bold'), fg='crimson').place(x=420, y=115+i*25)
			else:
				self.label = Label(self.parent, text=list_TA_predict[index_num][i], font=('Arial', 13), fg='darkslategrey').place(x=420, y=115+i*25)

		self.parent.protocol('WM_DELETE_WINDOW',self.close)

		menubar = Menu(self.parent)
		menu1 = Menu(menubar, tearoff=0)
		menu1.add_command(label="Про програму", command=lambda: self.goto_help(1))
		menu1.add_command(label="Технічний аналіз", command=lambda: self.goto_help(2))
		menu1.add_command(label="ARIMA", command=lambda: self.goto_help(3))
		menubar.add_cascade(label="Допомога", menu=menu1)
		menu0 = Menu(menubar, tearoff=0)
		menu0.add_command(label="Змінити параметри за замовчуванням (для "+names[index_num]+")", command=lambda: self.goto_change())
		menubar.add_cascade(label="Налаштування", menu=menu0)
		self.parent.config(menu=menubar)

	def goto_change(self):
		window = Toplevel(self.parent)
		my_window = change_window(window)

	def goto_help(self, b):
		global help_num
		help_num = b
		window = Toplevel(self.parent)
		my_window = help_window(window)

	def centerWindow(self):
		w = 480
		h = 490
		x = (1536 - w)/2
		y = (864 - h)/2
		self.parent.geometry('%dx%d+%d+%d' % (w, h, x, y))

	def quit(self):
		self.parent.destroy()

	def close(self):
		self.parent.grab_release()
		self.parent.destroy()

class help_window():
	def __init__(self, parent):
		self.parent = parent
		self.parent.title("Допомога")
		self.parent.iconbitmap(r'help.ico')
		self.parent.resizable(False, False)

		if help_num == 1:
			self.centerWindow(600, 300)
			self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 3)).pack()
			self.label0 = Label(self.parent, text="Про програму", anchor=N, font=('Arial', 20, 'bold'), fg='red').pack()
			self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 1)).pack()
			self.label1 = Label(self.parent, anchor=N, text='Програма дозволяє отримати сигнали для відкриття/закриття позиції', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='для 7 індексів за допомогою методів ТА та ARIMA-моделей. Дані', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='оновлюються з кожним запуском програми.\n', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='Користувач може змінити параметри методів за замовчуванням.\n', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='Якщо Ви виявили помилки в роботі програми або маєте пропозиції', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='щодо її покращення, повідомте про це victortaraba@gmail.com.', font=('Times New Roman', 14)).pack()			

		elif help_num == 2:
			self.centerWindow(600, 420)
			self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 3)).pack()
			self.label0 = Label(self.parent, text="Технічний аналіз", anchor=N, font=('Arial', 20, 'bold'), fg='red').pack()
			self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 1)).pack()
			self.label1 = Label(self.parent, anchor=N, text='У програмі реалізовано 13 індикаторів: перетин ціни з SMA, EMA,', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='LWMA, Aroon, CCI, SO, CMO, MAE(SMA,LL), MAE(SMA,UL),', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='MAE(EMA,LL), MAE(EMA,UL), MAE(LWMA,LL), MAE(LWMA,UL).\n', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='-1 - сигнал на відкриття короткої позиції та на закриття довгої;', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='1 - сигнал на відкриття довгої позиції та на закриття короткої;', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='0 - нічого не змінюємо.\n', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='Оптимальні параметри (параметри за замовчуванням) обрано наступним', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='чином: на основі даних 2010-2018 рр. для кожного індикатора для', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='кожного ринку обирається 5 результатів з найбільшою дохідністю; ці 5', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='індикаторів з тими ж параметрами тестуються на даних за 2018-2020 рр.,', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='серед них обирається той, який показує найкращі результати.\n', font=('Times New Roman', 14)).pack()

		elif help_num == 3:
			self.centerWindow(695, 555)
			self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 3)).pack()
			self.label0 = Label(self.parent, text="ARIMA", anchor=N, font=('Arial', 20, 'bold'), fg='red').pack()
			self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 1)).pack()
			self.label1 = Label(self.parent, anchor=N, text='Програма обирає оптимальну ARIMA-модель для 0<=p<=3, 0<=q<=3 (крім', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='того, порівнюються моделі з константою та без, та моделі з константою та', font=('Times New Roman', 14)).pack()	
			self.label1 = Label(self.parent, anchor=N, text='без, та моделі з параметрами, обраними за замовчуванням, та перевіряється', font=('Times New Roman', 14)).pack()
			self.label1 = Label(self.parent, anchor=N, text='значущість коефіцієнтів).', font=('Times New Roman', 14)).pack()

			photo01 = Image.open("ARMA.jpg")
			photo01 = photo01.resize((680, 140), Image.ANTIALIAS)
			photo1 = ImageTk.PhotoImage(photo01, master=self.parent)
			self.label_photo = Label(self.parent, image=photo1)
			self.label_photo.image = photo1
			self.label_photo.place(x=5, y=180)

			photo01 = Image.open("ARIMA.jpg")
			photo01 = photo01.resize((680, 145), Image.ANTIALIAS)
			photo1 = ImageTk.PhotoImage(photo01, master=self.parent)
			self.label_photo = Label(self.parent, image=photo1)
			self.label_photo.image = photo1
			self.label_photo.place(x=5, y=325)
			
			self.label1 = Label(self.parent, text='Якщо замість прогнозу відображається "---", то це означає, що програмі не вдалося', font=('Times New Roman', 14)).place(x=5, y=485)
			self.label = Label(self.parent, text='підібрати оптимальну модель.', font=('Times New Roman', 14)).place(x=220, y=515)

	def centerWindow(self, w, h):
		x = (1536 - w)/2
		y = (864 - h)/2
		self.parent.geometry('%dx%d+%d+%d' % (w, h, x, y))

class change_window():
	def __init__(self, parent):
		self.parent = parent
		self.parent.title("Налаштування")
		self.parent.iconbitmap(r'gear.ico')
		self.parent.resizable(False, False)
		self.set_window()

	def set_window(self):
		self.centerWindow(500, 475)
		self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 3)).pack()
		self.label0 = Label(self.parent, text="Зміна параметрів", anchor=N, font=('Arial', 20, 'bold'), fg='red').pack()
		self.label0 = Label(self.parent, text='\n', anchor=N, font=('Arial', 1)).pack()
		self.label1 = Label(self.parent, anchor=N, text='Оберіть метод, параметри якого бажаєте змінити:', font=('Times New Roman', 14)).pack()

		clicked = StringVar()
		clicked.set(TA_name[0])
		self.optionmenu = OptionMenu(self.parent, clicked, *TA_name)
		self.optionmenu.pack()
		self.optionmenu.config(font=('Arial', 10))
		self.menu = self.optionmenu.nametowidget(self.optionmenu.menuname)
		self.menu.config(font=('Arial', 10))

		self.label1 = Label(self.parent, anchor=N, text='\n'*5).pack()
		self.label1 = Label(self.parent, anchor=N, text='Увага!', font=('Times New Roman', 18, 'bold'), fg='red').pack()
		self.label1 = Label(self.parent, anchor=N, text='Відновити початкові параметри після зміни', font=('Times New Roman', 14)).pack()
		self.label1 = Label(self.parent, anchor=N, text='неможливо', font=('Times New Roman', 14)).pack()
		self.label1 = Label(self.parent, anchor=N, text='\n'*4).pack()
		self.button1 = Button(self.parent, anchor=N, text="Підтвердити вибір\nі перейти до вибору параметрів", font = ('Arial', 14, 'bold'), activebackground='blue', activeforeground='white',\
			fg='blue', bg='white', width=35, relief=GROOVE, command=lambda: self.goto_change_param(TA_name.index(clicked.get()))).pack()

	def goto_change_param(self, c):
		global name_ind_param
		name_ind_param = c
		window = Toplevel(self.parent)
		my_window = change_param_window(window)
		self.parent.withdraw()

	def centerWindow(self, w, h):
		x = (1536 - w)/2
		y = (864 - h)/2
		self.parent.geometry('%dx%d+%d+%d' % (w, h, x, y))

class change_param_window():
	def __init__(self, parent):
		self.parent = parent
		self.parent.title("Зміна параметрів за замовчуванням")
		self.parent.iconbitmap(r'gear.ico')
		self.parent.resizable(False, False)
		self.c_num = index_num+1
		self.r_num = name_ind_param+1
		self.set_window()

	def set_window(self):
		if self.r_num>=8:
			self.centerWindow(460, 370)
			self.label0 = Label(self.parent, text='\n', font=('Arial', 3)).pack()
			self.label0 = Label(self.parent, text="Зміна параметрів", anchor=N, font=('Arial', 20, 'bold'), fg='red').pack()
			self.label0 = Label(self.parent, text='\n', font=('Arial', 1)).pack()
			self.label0 = Label(self.parent, anchor=N, text='Індекс: '+names[self.c_num-1]+'. Метод: '+TA_name[self.r_num-1]+'\n', font=('Times New Roman', 14), fg='blue').pack()
			self.label0 = Label(self.parent, anchor=N, text='Увага!', font=('Times New Roman', 14, 'bold'), fg='red').pack()
			self.label0 = Label(self.parent, anchor=N, text='Формат введення: два значення через пробіл', font=('Times New Roman', 14), fg='red').pack()
			self.label0 = Label(self.parent, anchor=N, text='(p - відсоток для верхньої/нижньої межі, n)\n', font=('Times New Roman', 14), fg='red').pack()
			self.label0 = Label(self.parent, anchor=N, text='Введіть нові значення параметрів:', font=('Times New Roman', 14)).pack()
			self.label_incorrect = Label(self.parent, anchor=N, text='\n', font=('Times New Roman', 1)).pack()
			self.n = Entry(self.parent)
			self.n.pack()
			self.label0 = Label(self.parent, text='\n', font=('Arial', 2)).pack()
			self.label_incorrect = Label(self.parent, text='\n', font=('Times New Roman', 14, 'bold'), fg='red')
			self.label_incorrect.pack()
			self.button = Button(self.parent, anchor=N, text="Підтвердити", font = ('Times New Roman', 14, 'bold'), activebackground='blue', activeforeground='white',\
				fg='blue', bg='white', width=15, relief=GROOVE, command=self.final_button_2)
			self.button.place(x=145, y=315)
		else:
			self.centerWindow(350, 280)
			self.label0 = Label(self.parent, text='\n', font=('Arial', 3)).pack()
			self.label0 = Label(self.parent, text="Зміна параметрів", anchor=N, font=('Arial', 20, 'bold'), fg='red').pack()
			self.label0 = Label(self.parent, text='\n', font=('Arial', 1)).pack()
			self.label0 = Label(self.parent, anchor=N, text='Індекс: '+names[self.c_num-1]+'. Метод: '+TA_name[self.r_num-1]+'\n', font=('Times New Roman', 14), fg='blue').pack()
			self.label0 = Label(self.parent, anchor=N, text='Введіть нове значення параметра n:', font=('Times New Roman', 14)).pack()
			self.label_incorrect = Label(self.parent, anchor=N, text='\n', font=('Times New Roman', 1)).pack()
			self.n = Entry(self.parent)
			self.n.pack()
			self.label0 = Label(self.parent, text='\n', font=('Arial', 2)).pack()
			self.label_incorrect = Label(self.parent, text='\n', font=('Times New Roman', 14, 'bold'), fg='red')
			self.label_incorrect.pack()
			self.button = Button(self.parent, anchor=N, text="Підтвердити", font = ('Times New Roman', 14, 'bold'), activebackground='blue', activeforeground='white',\
				fg='blue', bg='white', width=15, relief=GROOVE, command=self.final_button_1).pack()

	def final_button_1(self):
		try:
			input_ = int(self.n.get())
			if input_ > 100 or input_ <=1:
				raise ValueError
			self.label_incorrect["text"] = ""
			
			workbook = load_workbook(filename="C_TA_param.xlsx")
			sheet = workbook.active
			sheet.cell(row=self.r_num, column=self.c_num).value = input_
			workbook.save(filename="C_TA_param.xlsx")
			workbook.close()

			messagebox.showinfo("Зміна параметрів","Запис в БД змінено. Будь ласка, перезапустіть програму")
			self.parent.destroy()

		except ValueError:
			self.label_incorrect["text"] = "Введіть число, яке > 1, і <= 100"

	def final_button_2(self):
		try:
			text = self.n.get()
			text = text.split(' ')

			if len(text) != 2:
				raise ValueError
			input_per = float(text[0])
			input_n = int(text[1])
			if input_per > 1 or input_per <=0:
				raise ValueError
			if input_n > 100 or input_n <=1:
				raise ValueError
			self.label_incorrect["text"] = ""
			
			workbook = load_workbook(filename="C_TA_param.xlsx")
			sheet = workbook.active

			r_num_1 = self.r_num + (self.r_num-8)
			r_num_2 = r_num_1 + 1
			sheet.cell(row=r_num_1, column=self.c_num).value = input_per
			sheet.cell(row=r_num_2, column=self.c_num).value = input_n

			workbook.save(filename="C_TA_param.xlsx")
			workbook.close()

			messagebox.showinfo("Зміна параметрів","Запис в БД змінено. Будь ласка, перезапустіть програму")
			self.parent.destroy()

		except ValueError:
			self.label_incorrect["text"] = "Допустимі значення параметрів: 1<n<=100, 0<p<=1"

	def centerWindow(self, w, h):
		x = (1536 - w)/2
		y = (864 - h)/2
		self.parent.geometry('%dx%d+%d+%d' % (w, h, x, y))

def main():
	root = Tk()
	app = Invest_model(root)
	root.resizable(False, False)
	root.mainloop()

main()