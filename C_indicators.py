import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from C_download_data import index_list
from openpyxl import load_workbook

def SMA(data,n):
	sma_list=[np.nan for k in range(n-1)]
	for i in range(len(data)-n+1):
		sum_n = 0
		for j in range(n):
			sum_n = sum_n + data[i+j]
		sma=sum_n/n
		sma_list.append(sma)
	return sma_list

def LWMA(data,n):
	lma_list=[np.nan for k in range(n-1)]

	for i in range(len(data)-n+1):
		sum_n, sum_j = 0, 0
		for j in range(n):
			sum_n = sum_n + (j+1)*data[i+j]
			sum_j = sum_j + (j+1)
		lma=sum_n/sum_j
		lma_list.append(lma)

	return lma_list

def EMA(data, n, a=0):
	ema_list=[data[0]]
	if a==0:
		a=2/(n+1)

	for i in range (1,len(data)):
		ema = 0
		ema = (1-a)*ema_list[i-1]+a*data[i]
		ema_list.append(ema)

	return ema_list

def MACD(data, n1, n2, n3):
	EMA_n1 = EMA(data,n1)
	EMA_n2 = EMA(data,n2)
	macd_list, h_list = [], []

	for i in range(len(data)):
		macd_list.append(EMA_n1[i]-EMA_n2[i])
	EMA_n3 = EMA(macd_list,n3) 
	for i in range(len(data)):
		h_list.append(macd_list[i]-EMA_n3[i])

	return h_list

def RSI(data, n):
	U, D = [np.nan], [np.nan]

	for i in range(1,len(data)):
		if data[i]>data[i-1]:
			U.append(data[i]-data[i-1])
			D.append(0)
		elif data[i]<data[i-1]:
			U.append(0)
			D.append(data[i-1]-data[i])
		else:
			U.append(0)
			D.append(0)

	SMA_U = SMA(U, n)
	SMA_D = SMA(D, n)

	RS_list = []
	for i in range(len(data)):
		if SMA_D[i] != 0:
			RS_list.append(SMA_U[i]/SMA_D[i])
		else:
			RS_list.append("RSI=100")

	RSI_list=[]
	for i in range(len(RS_list)):
		if RS_list[i] != "RSI=100":
			RSI_list.append(100*RS_list[i]/(1+RS_list[i]))
		else:
			RSI_list.append(100)

	return RSI_list

def Aroon(data, n):
	Up, Dn = [np.nan for k in range(n-1)], [np.nan for k in range(n-1)]

	for i in range(len(data)-n+1):
		list_n = []
		for j in range(0, n):
			list_n.append(data[i+j])
		
		min_index = list_n.index(min(list_n))+1 
		max_index = list_n.index(max(list_n))+1
		
		Up.append((n-(n-max_index))*100/n)
		Dn.append((n-(n-min_index))*100/n)
	return Up, Dn

def CCI(data_close, data_high, data_low, n):
	p_typical = []

	for i in range(len(data_close)):
		p_typical.append((data_close[i]+data_low[i]+data_high[i])/3)
	sma = SMA(p_typical,n)

	MAD = [np.nan for k in range(n-1)]
	for j in range(n-1, len(data_close)):
		mad = 0
		for s in range(n):
			mad = mad +  abs(p_typical[j-s]-sma[j])
		MAD.append(mad/n)

	CCI = [np.nan for n in range(n-1)]
	for t in range(n-1, len(data_close)):
		cci = (p_typical[t]-sma[t])/(0.015*MAD[t])
		CCI.append(cci)

	return CCI

def KAMA(data, n, n1, n2):

	ABS_n_periods_ago = [np.nan for k in range(n-1)]
	for i in range(n-1, len(data)):
		ABS_n_periods_ago.append(abs(data[i]-data[i-n]))

	ABS_prior_close = [np.nan]
	for i in range(1, len(data)):
		ABS_prior_close.append(abs(data[i]-data[i-1]))

	ER = [np.nan for k in range(n-1)]
	for i in range(n-1, len(data)):
		numerator = ABS_n_periods_ago[i]
		denominator = 0
		for s in range(n):
			denominator = denominator + ABS_prior_close[i-s]
		ER.append(numerator/denominator)

	a_fast = 2/(1+n1)
	a_slow = 2/(1+n2)

	SC = [np.nan for k in range(n-1)]
	for i in range(n-1, len(data)):
		SC.append(((a_fast-a_slow)*ER[i]+a_slow)**2)

	KAMA = [np.nan for k in range(n-1)]
	KAMA.append(data[n-1])
	for i in range(n, len(data)):
		KAMA.append(KAMA[i-1]+SC[i]*(data[i]-KAMA[i-1]))

	return KAMA

def SO(data_close, data_high, data_low, n):
	K, D = [np.nan for k in range(n-1)], [np.nan for k in range(n-1)]
	highest_high, lowest_low = [np.nan for k in range(n-1)], [np.nan for k in range(n-1)]

	for i in range(len(data_close)-n+1):
		list_n_high, list_n_low = [], []
		for j in range(0, n):
			list_n_high.append(data_high[i+j])
			list_n_low.append(data_low[i+j])
		highest_high.append(max(list_n_high))
		lowest_low.append(min(list_n_low))

	for i in range(n-1, len(highest_high)):
		K.append(100*(data_close[i]-lowest_low[i])/(highest_high[i]-lowest_low[i]))
	D = SMA(K, n)

	return K, D

def CHO(data_close, data_low, data_high, data_volume, n1, n2):
	CMFV=[]
	for i in range(len(data_close)):
		CMFV.append(((data_close[i]-data_low[i])-(data_high[i]-data_close[i]))*data_volume[i]/(data_high[i]-data_low[i]))

	AD=[CMFV[0]]
	for i in range(1, len(CMFV)):
		ad = AD[i-1]+CMFV[i]
		AD.append(ad)

	MA_n1 = EMA(AD, n1)
	MA_n2 = EMA(AD, n2)
	CHO_list=[]
	for i in range(len(MA_n1)):
		CHO_list.append(MA_n1[i]-MA_n2[i])

	return CHO_list

def CMO(data, n):
	CMO1, CMO2 = [np.nan], [np.nan]

	for i in range(1,len(data)):
		if data[i]>data[i-1]:
			CMO1.append(data[i]-data[i-1])
			CMO2.append(0)
		elif data[i]<data[i-1]:
			CMO1.append(0)
			CMO2.append(data[i-1]-data[i])
		else:
			CMO1.append(0)
			CMO2.append(0)

	sH, sL = [np.nan for k in range(n)], [np.nan for k in range(n)]
	for i in range(n,len(data)):
		sh, sl = 0, 0
		for t in range(n):
			sh = sh + CMO1[i-t]
			sl = sl + CMO2[i-t]
		sH.append(sh)
		sL.append(sl)

	CMO_list=[np.nan for k in range(n)] 
	for i in range(n,len(sH)):
		if (sH[i]+sL[i]) != 0:
			CMO_list.append(100*(sH[i]-sL[i])/(sH[i]+sL[i]))
		else:
			CMO_list.append(100*(sH[i-1]-sL[i-1])/(sH[i-1]+sL[i-1]))
		

	return CMO_list

def strategy_MA_1(P, n1, MA_type, n2=26, n3=9):
	MA= []
	if MA_type == "SMA":
		MA = SMA(P, n1)
	elif MA_type == "EMA":
		MA = EMA(P, n1)
	elif MA_type == "LWMA":
		MA = LWMA(P, n1)
	elif MA_type == "KAMA":
		MA = KAMA(P, n1, n2, n3)

	if MA_type != "EMA":
		s=[np.nan for k in range(n1)]
		for i in range(n1, len(MA)):
			if (P[i-1]<MA[i-1]) and (P[i]>MA[i]):
				s.append(1)
			elif (P[i-1]>MA[i-1]) and (P[i]<MA[i]):
				s.append(-1)
			else:
				s.append(0)
	elif MA_type == "EMA":
		s=[np.nan]
		for i in range(1,len(MA)):
			if (P[i-1]<MA[i-1]) and (P[i]>MA[i]):
				s.append(1)
			elif (P[i-1]>MA[i-1]) and (P[i]<MA[i]):
				s.append(-1)
			else:
				s.append(0)
	return s

def strategy_RSI(P, n):
	RSI_list = RSI(P, n)
	s=[]
	for i in range(len(P)):
		if (RSI_list[i]>70) and (RSI_list[i-1]<70):
			s.append(1)
		elif (RSI_list[i]<70) and (RSI_list[i-1]>70):
			s.append(-1)
		else:
			s.append(0)
	return s

def strategy_CMO(P, n):
	CMO_list = CMO(P, n)
	s=[np.nan for k in range(n)]
	for i in range(n, len(P)):
		if (CMO_list[i]>0) and (CMO_list[i-1]<0):
			s.append(1)
		elif (CMO_list[i]<0) and (CMO_list[i-1]>0):
			s.append(-1)
		else:
			s.append(0)
	return s

def strategy_Aroon(P,n):
	up, dn = Aroon(P, n)
	s=[]
	for i in range(len(P)):
		if (up[i]>dn[i]) and (up[i-1]<dn[i-1]):
			s.append(1)
		elif (up[i]<dn[i]) and (up[i-1]>dn[i-1]):
			s.append(-1)
		else:
			s.append(0)
	return s

def strategy_SO(P_close, P_high, P_low, n):
	K, D = SO(P_close, P_high, P_low, n)
	s=[np.nan for k in range(n)]
	for i in range(n, len(P_close)):
		if (K[i]>D[i]) and (K[i-1]<D[i-1]):
			s.append(1)
		elif (K[i]<D[i]) and (K[i-1]>D[i-1]):
			s.append(-1)
		else:
			s.append(0)
	return s

def strategy_CCI(data_close, data_high, data_low, n):
	cci = CCI(data_close, data_high, data_low, n)
	s=[]
	for i in range(len(data_close)):
		if ((cci[i]>0) and (cci[i-1]<0)):
			s.append(1)
		elif ((cci[i]<0) and (cci[i-1]>0)):
			s.append(-1)
		else:
			s.append(0)
	return s

def strategy_CHO(data_close, data_low, data_high, data_volume, n1, n2):
	cho = CHO(data_close, data_low, data_high, data_volume, n1, n2)
	s=[]
	for i in range(len(data_close)):
		if (cho[i]>0) and (cho[i-1]<0):
			s.append(1)
		elif (cho[i]<0) and (cho[i-1]>0):
			s.append(-1)
		else:
			s.append(0)
	return s

def strategy_MAE(P, upper, lower, n1, MA_type, MAE_type="LL"):
	if n1 >= 0:
		UL, LL, MA = [], [], []
		if MA_type == "SMA":
			MA = SMA(P, n1)
		elif MA_type == "EMA":
			MA = EMA(P, n1)
		elif MA_type == "LWMA":
			MA = LWMA(P, n1)

		for i in range(len(P)):
			UL.append((1+upper)*MA[i])
			LL.append((1-lower)*MA[i])
		if MA_type != "EMA":
			s=[np.nan for k in range(n1)]
			if MAE_type == "UL":
				for i in range(n1, len(MA)):
					if (P[i-1]<UL[i-1]) and (P[i]>UL[i]):
						s.append(1)
					elif (P[i-1]>UL[i-1]) and (P[i]<UL[i]):
						s.append(-1)
					else:
						s.append(0)
			elif MAE_type == "LL":
				for i in range(n1, len(MA)):
					if (P[i-1]<LL[i-1]) and (P[i]>LL[i]):
						s.append(1)
					elif (P[i-1]>LL[i-1]) and (P[i]<LL[i]):
						s.append(-1)
					else:
						s.append(0)
		elif MA_type == "EMA":
			s=[np.nan]
			if MAE_type == "UL":
				for i in range(1, len(MA)):
					if (P[i-1]<UL[i-1]) and (P[i]>UL[i]):
						s.append(1)
					elif (P[i-1]>UL[i-1]) and (P[i]<UL[i]):
						s.append(-1)
					else:
						s.append(0)
			elif MAE_type == "LL":
				for i in range(1, len(MA)):
					if (P[i-1]<LL[i-1]) and (P[i]>LL[i]):
						s.append(1)
					elif (P[i-1]>LL[i-1]) and (P[i]<LL[i]):
						s.append(-1)
					else:
						s.append(0)
		return s
	else:
		return ["nan", "nan"]

def test(P, S):
	result = []
	if S[0] != "nan":
		i_start, i_end = "NaN", "NaN"
		#для коротких позицій
		for i in range(len(S)):
			result.append("NaN")
			if S[i] == -1:
				i_start = i
			if S[i] == 1:
				i_end = i
				if i_start == "NaN":
					i_end = "NaN"
				if i_end != "NaN":
					yield_ = -(P[i_end]-P[i_start])/P[i_start]
					result[i] = yield_

		i_start, i_end = "NaN", "NaN"
		#для довгих позицій
		for i in range(len(S)):
			if S[i] == 1:
				i_start = i
			if S[i] == -1:
				i_end = i
				if i_start == "NaN":
					i_end = "NaN"
				if i_end != "NaN":
					yield_ = (P[i_end]-P[i_start])/P[i_start]
					result[i] = yield_
		m = 100
		for i in range(len(result)):
			if result[i]!="NaN":
				m = m*(1+result[i])
				result[i] = m
		return m
	else:
		return -100

workbook = load_workbook(filename="C_TA_param.xlsx")
sheet = workbook.active
TA_param = [[] for i in range(7)]
for j in range(1,8):
	for i in range(1,20):
		TA_param[j-1].append(sheet.cell(row=i,column=j).value)
workbook.close()

list_TA_predict = [[] for i in range(7)]

for i in range(len(index_list)):
	index_TA_all_results = []
	data_index = index_list[i]['Close']

	index_TA_all_results.append(strategy_MA_1(data_index, TA_param[i][0], "SMA"))
	index_TA_all_results.append(strategy_MA_1(data_index, TA_param[i][1], "EMA"))
	index_TA_all_results.append(strategy_MA_1(data_index, TA_param[i][2], "LWMA"))
	index_TA_all_results.append(strategy_Aroon(data_index, TA_param[i][3]))
	index_TA_all_results.append(strategy_CCI(data_index, index_list[i]['High'], index_list[i]['Low'], TA_param[i][4]))
	index_TA_all_results.append(strategy_SO(data_index, index_list[i]['High'], index_list[i]['Low'], TA_param[i][5]))
	index_TA_all_results.append(strategy_CMO(data_index, TA_param[i][6]))
	index_TA_all_results.append(strategy_MAE(data_index, upper=0, lower=TA_param[i][7], n1=TA_param[i][8], MA_type="SMA", MAE_type="LL"))
	index_TA_all_results.append(strategy_MAE(data_index, upper=TA_param[i][9], lower=0, n1=TA_param[i][10], MA_type="SMA", MAE_type="UL"))
	index_TA_all_results.append(strategy_MAE(data_index, upper=0, lower=TA_param[i][11], n1=TA_param[i][12], MA_type="LWMA", MAE_type="LL"))
	index_TA_all_results.append(strategy_MAE(data_index, upper=TA_param[i][13], lower=0, n1=TA_param[i][14], MA_type="LWMA", MAE_type="UL"))
	index_TA_all_results.append(strategy_MAE(data_index, upper=0, lower=TA_param[i][15], n1=TA_param[i][16], MA_type="EMA", MAE_type="LL"))
	index_TA_all_results.append(strategy_MAE(data_index, upper=TA_param[i][17], lower=0, n1=TA_param[i][18], MA_type="EMA", MAE_type="UL"))

	for j in range(len(index_TA_all_results)):
		list_TA_predict[i].append(index_TA_all_results[j][len(index_TA_all_results[j])-1])